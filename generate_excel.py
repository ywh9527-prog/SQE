import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# 外协交付业绩数据
waixie_data = [
    ["厂商", "1月", "2月", "3月", "4月", "5月", "6月", "7月", "8月", "9月", "10月", "11月", "12月"],
    ["华弘鑫", "", 60.00, 60.00, 84.29, 95.00, 82.98, 88.13, 88.13, 88.13, 96.76, 88.30, 88.30],
    ["惠开", "", "", "", "", "", 94.60, 85.95, 85.95, 85.95, 87.43, 88.33, 88.33],
    ["晶蓝", 51.21, "", 73.33, 94.29, 80.00, 51.00, 80.00, 80.00, 80.00, 67.86, 75.33, 85.33],
    ["晶星", "", "", "", "", "", "", "", "", "", "", "", 92.50],
    ["辽源", "", "", 100.00, 95.00, 86.67, 88.00, 80.00, 80.00, 80.00, 90.00, 90.00, ""],
    ["锐盛", "", "", "", "", "", 60.00, 100.00, 100.00, 100.00, 80.00, 80.00, ""],
    ["盛宥", "", "", 60.00, 90.00, 100.00, 100.00, 80.00, 100.00, "", 81.43, 84.00, 84.00],
    ["祥威", "", "", "", "", "", "", "", "", "", "", "", 89.09],
    ["新显", "", "", "", "", "", 70.91, 62.22, 62.22, 62.22, 64.21, 64.21, 100.00],
    ["幸福电子", "", "", "", 100.00, 100.00, 100.00, 92.26, 92.26, 82.26, 97.14, 97.14, ""],
    ["旭虹", "", 75.45, 65.75, 83.85, 77.14, 89.83, 60.00, 60.00, 60.00, 88.57, 100.00, 100.00],
    ["宣益", "", "", "", "", "", "", "", "", "", "", 76.00, 76.00],
    ["亦高", "", "", "", "", "", "", "", "", "", 62.96, 62.96, 100.00],
    ["源合达", "", "", "", "", "", 60.00, 60.00, 60.00, 80.00, 80.00, 80.00, ""],
    ["中科", "", "", "", "", "", "", "", "", "", "", 92.50, 92.50],
    ["仁轩", 75.80, 58.50, 98.00, "", "", 80.00, 100.00, 100.00, 84.78, 90.00, 84.67, 84.67]
]

# 外购交付业绩数据
waigou_data = [
    ["供应商", "1月", "2月", "3月", "4月", "5月", "6月", "7月", "8月", "9月", "10月", "11月", "12月"],
    ["虹科", "", "", "", "", "", 100.00, 100.00, 100.00, 100.00, "", 100.00, 100.00],
    ["格莱特", "", "", "", "", "", "", 100.00, 100.00, 100.00, 100.00, "", 100.00],
    ["森耐", "", "", "", "", "", "", 100.00, 100.00, 100.00, 100.00, 100.00, 100.00],
    ["宇皓", "", "", "", "", "", "", 100.00, 100.00, 100.00, 100.00, 100.00, 100.00],
    ["BKTECH CO.,LTD", 100.00, 100.00, 100.00, 100.00, 100.00, 100.00, 100.00, 100.00, 100.00, 100.00, 92.00, 92.00],
    ["百川兴", 100.00, 100.00, 92.50, 87.73, 100.00, 100.00, 100.00, 100.00, 100.00, 100.00, 94.29, 94.29],
    ["创新", 86.67, 86.67, 100.00, 60.00, 60.00, 100.00, 100.00, 100.00, 100.00, 100.00, 100.00, 100.00],
    ["鑫诚", 80.00, 100.00, 100.00, 100.00, 95.00, 93.00, 100.00, 100.00, 100.00, 100.00, 100.00, 100.00],
    ["弘擎", "", "", "", "", "", "", 100.00, 100.00, 100.00, 100.00, 100.00, 100.00],
    ["利丰达", 100.00, "", 100.00, 100.00, "", 100.00, 100.00, 100.00, 100.00, 100.00, 100.00, 100.00],
    ["旭虹（ASF）", 100.00, "", 100.00, 100.00, "", 100.00, 100.00, 100.00, 100.00, 100.00, 100.00, 100.00],
    ["触尔发", 90.00, 81.00, 100.00, 100.00, 98.00, 99.00, 100.00, 100.00, 85.00, 90.00, 100.00, 100.00],
    ["舜利", 100.00, 100.00, 100.00, 100.00, 100.00, 100.00, 98.62, 98.62, 88.62, 95.00, 94.23, 95.00],
    ["旭虹（1:1）", 86.67, 100.00, 100.00, 100.00, 100.00, 100.00, 100.00, 100.00, 100.00, 100.00, 100.00, 100.00],
    ["致贯", "", "", "", "", "", "", 100.00, 100.00, 100.00, 100.00, 100.00, 100.00],
    ["点赞", "", "", "", "", "", "", 100.00, 100.00, 100.00, 100.00, 100.00, 100.00],
    ["富行", 100.00, "", 100.00, 95.00, 100.00, 100.00, 100.00, 100.00, 100.00, 100.00, 100.00, 100.00],
    ["龙源", "", "", "", "", 100.00, 100.00, 100.00, 100.00, 100.00, 100.00, 100.00, 100.00],
    ["常兴", 100.00, 98.26, 100.00, 100.00, 100.00, 100.00, 100.00, 100.00, 100.00, 100.00, 96.00, 100.00],
    ["索威斯", "", "", "", "", "", "", 100.00, 100.00, "", "", 100.00, ""],
    ["旭航", "", "", "", "", "", "", 100.00, 100.00, 100.00, 100.00, 100.00, 100.00],
    ["一米新", "", "", "", "", "", "", 100.00, 100.00, 100.00, 100.00, 100.00, 100.00]
]

# 创建Excel工作簿
wb = Workbook()

# 样式定义
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
header_alignment = Alignment(horizontal="center", vertical="center")
header_border = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000")
)

data_font = Font(name="微软雅黑", size=10)
data_alignment = Alignment(horizontal="center", vertical="center")
data_border = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000")
)

# 创建外协表格
ws1 = wb.active
ws1.title = "外协交付业绩"

for row_idx, row_data in enumerate(waixie_data, start=1):
    for col_idx, cell_value in enumerate(row_data, start=1):
        cell = ws1.cell(row=row_idx, column=col_idx, value=cell_value)

        if row_idx == 1:  # 标题行
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = header_border
        else:  # 数据行
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = data_border

# 设置列宽
ws1.column_dimensions['A'].width = 15
for col in range(2, 14):
    ws1.column_dimensions[get_column_letter(col)].width = 8

# 创建外购表格
ws2 = wb.create_sheet("外购交付业绩")

for row_idx, row_data in enumerate(waigou_data, start=1):
    for col_idx, cell_value in enumerate(row_data, start=1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=cell_value)

        if row_idx == 1:  # 标题行
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = header_border
        else:  # 数据行
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = data_border

# 设置列宽
ws2.column_dimensions['A'].width = 18
for col in range(2, 14):
    ws2.column_dimensions[get_column_letter(col)].width = 8

# 保存文件
output_file = r"D:\AI\Claude-SQE-Data-Analysis-Assistant-refactored\2025年交付业绩汇总.xlsx"
wb.save(output_file)
print(f"Excel文件已生成: {output_file}")